"""Pipeline module for model: petani-it6787-kampar-karsem-01.

Step 1: download data from Google Drive based on models/models.csv.
Step 2: generate output sheets (metadata, unique_farmer, unique_land_parcel, training).
"""

from __future__ import annotations

import csv
import os
import re
import configparser
from typing import Dict, List, Tuple

import gdown
import pandas as pd
import requests
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

DEFAULT_MODEL_CSV = os.path.join("models", "models.csv")


class ModelConfigNotFoundError(ValueError):
    """Raised when a model_id is missing from models.csv."""


def _extract_gdrive_id(source: str) -> str:
    """Extract Google Drive file id from a URL or return the id as-is."""
    source = source.strip()
    if "drive.google.com" not in source:
        return source

    # Common patterns: /file/d/<id>/ or id=<id>
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", source)
    if match:
        return match.group(1)

    match = re.search(r"id=([a-zA-Z0-9_-]+)", source)
    if match:
        return match.group(1)

    # Fallback: last path segment
    parts = [p for p in source.split("/") if p]
    return parts[-1]


def _load_model_config(model_id: str, csv_path: str = DEFAULT_MODEL_CSV) -> Dict[str, str]:
    """Load a single model config row by model_id from CSV."""
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("model_id") or "").strip() == model_id:
                # Normalize keys by stripping whitespace
                return {k.strip(): (v or "").strip() for k, v in row.items()}

    raise ModelConfigNotFoundError(f"model_id not found in {csv_path}: {model_id}")


def _detect_source_type(source: str) -> str:
    """Return 'gspread', 'excel', or 'unknown' based on source URL or id."""
    s = source.lower().strip()
    if "docs.google.com/spreadsheets" in s or "spreadsheets/d/" in s:
        return "gspread"
    if s.endswith(".xlsx") or "drive.google.com" in s:
        return "excel"
    return "unknown"


def _build_output_path(cfg: Dict[str, str]) -> str:
    input_folder = cfg.get("input_folder", "")
    input_name = cfg.get("input_name", "")
    if not input_folder or not input_name:
        raise ValueError("input_folder/input_name is empty for model_id: " + cfg.get("model_id", ""))
    os.makedirs(input_folder, exist_ok=True)
    return os.path.join(input_folder, f"{input_name}.xlsx")


def _build_output_file_path(cfg: Dict[str, str]) -> str:
    output_folder = cfg.get("output_folder", "")
    output_name = cfg.get("output_name", "")
    if not output_folder or not output_name:
        raise ValueError("output_folder/output_name is empty for model_id: " + cfg.get("model_id", ""))
    os.makedirs(output_folder, exist_ok=True)
    return os.path.join(output_folder, f"{output_name}.xlsx")


def _download_google_sheet(file_id: str, output_path: str) -> None:
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    resp = requests.get(url, stream=True, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"Failed export Google Sheet: HTTP {resp.status_code}")
    with open(output_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=1024 * 1024):
            if chunk:
                f.write(chunk)


def _download_gdrive_file(file_id: str, output_path: str) -> None:
    gdown.download(id=file_id, output=output_path, quiet=False)


def _get_sheet_list(path: str) -> List[str]:
    xls = pd.ExcelFile(path)
    return list(xls.sheet_names)


def _col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    idx = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {col}")
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _resolve_params_path(cfg: Dict[str, str]) -> str:
    params_file = cfg.get("params", "").strip()
    if not params_file:
        raise ValueError("params is empty for model_id: " + cfg.get("model_id", ""))
    if os.path.isabs(params_file):
        return params_file
    return os.path.join(os.path.dirname(__file__), params_file)


def _load_params(cfg: Dict[str, str]) -> Dict[str, Dict[str, str]]:
    path = _resolve_params_path(cfg)
    if not os.path.exists(path):
        raise FileNotFoundError(f"params file not found: {path}")
    parser = configparser.ConfigParser()
    parser.read(path)
    return {section: dict(parser.items(section)) for section in parser.sections()}


def _extract_metadata_airterbit(path: str, params: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    meta = params.get("metadata", {})
    sheet_name = meta.get("sheet_name", "AIR TERBIT")
    row_start = int(meta.get("row_start", "2"))
    row_end = int(meta.get("row_end", "6"))
    col_label = meta.get("col_label", "A")
    col_value = meta.get("col_value", "D")

    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    r0 = row_start - 1
    r1 = row_end
    c_label = _col_letter_to_index(col_label)
    c_value = _col_letter_to_index(col_value)

    labels = df.iloc[r0:r1, c_label].astype(str).tolist()
    values = df.iloc[r0:r1, c_value].astype(str).tolist()

    return pd.DataFrame({"label": labels, "value": values})


def _extract_unique_farmer(path: str, params: Dict[str, Dict[str, str]], section: str) -> pd.DataFrame:
    cfg = params.get(section, {})
    sheet_name = cfg.get("sheet_name", "")
    row_start = int(cfg.get("row_start", "1"))
    row_end = int(cfg.get("row_end", "1"))
    col_id_petani = cfg.get("col_id_petani", "")
    col_nama_petani = cfg.get("col_nama_petani", "")
    col_nik = cfg.get("col_nik", "")
    col_jenis_kelamin = cfg.get("col_jenis_kelamin", "")

    if not sheet_name:
        raise ValueError(f"sheet_name missing in params section: {section}")
    if not col_id_petani or not col_nama_petani or not col_nik or not col_jenis_kelamin:
        raise ValueError(f"column mapping missing in params section: {section}")

    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    r0 = row_start - 1
    r1 = row_end
    c_id = _col_letter_to_index(col_id_petani)
    c_nama = _col_letter_to_index(col_nama_petani)
    c_nik = _col_letter_to_index(col_nik)
    c_gender = _col_letter_to_index(col_jenis_kelamin)

    out = pd.DataFrame(
        {
            "ID Petani": df.iloc[r0:r1, c_id].astype(str),
            "Nama Petani": df.iloc[r0:r1, c_nama].astype(str),
            "NIK": df.iloc[r0:r1, c_nik].astype(str),
            "Jenis Kelamin": df.iloc[r0:r1, c_gender].astype(str),
        }
    )

    out = out[out["ID Petani"].str.strip() != ""]
    out = out.drop_duplicates(subset=["ID Petani"])
    return out


def _extract_unique_land_parcel(path: str, params: Dict[str, Dict[str, str]], section: str) -> pd.DataFrame:
    cfg = params.get(section, {})
    sheet_name = cfg.get("sheet_name", "")
    row_start = int(cfg.get("row_start", "1"))
    row_end = int(cfg.get("row_end", "1"))
    col_id_petani = cfg.get("col_id_petani", "")
    col_id_lahan = cfg.get("col_id_lahan", "")
    col_nama_petani = cfg.get("col_nama_petani", "")
    col_nik = cfg.get("col_nik", "")
    col_jenis_kelamin = cfg.get("col_jenis_kelamin", "")

    if not sheet_name:
        raise ValueError(f"sheet_name missing in params section: {section}")
    if not col_id_petani or not col_id_lahan:
        raise ValueError(f"column mapping missing in params section: {section}")

    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    r0 = row_start - 1
    r1 = row_end
    c_id_petani = _col_letter_to_index(col_id_petani)
    c_id_lahan = _col_letter_to_index(col_id_lahan)
    c_nama = _col_letter_to_index(col_nama_petani) if col_nama_petani else None
    c_nik = _col_letter_to_index(col_nik) if col_nik else None
    c_gender = _col_letter_to_index(col_jenis_kelamin) if col_jenis_kelamin else None

    out = pd.DataFrame(
        {
            "ID Lahan": df.iloc[r0:r1, c_id_lahan].astype(str),
            "ID Petani": df.iloc[r0:r1, c_id_petani].astype(str),
            "Nama Petani": df.iloc[r0:r1, c_nama].astype(str) if c_nama is not None else "",
            "NIK": df.iloc[r0:r1, c_nik].astype(str) if c_nik is not None else "",
            "Jenis Kelamin": df.iloc[r0:r1, c_gender].astype(str) if c_gender is not None else "",
        }
    )

    out = out[out["ID Lahan"].str.strip() != ""]
    out = out.drop_duplicates(subset=["ID Lahan"])
    return out


def _extract_training(path: str, params: Dict[str, Dict[str, str]], section: str) -> pd.DataFrame:
    cfg = params.get(section, {})
    sheet_name = cfg.get("sheet_name", "")
    row_start = int(cfg.get("row_start", "1"))
    row_end = int(cfg.get("row_end", "1"))
    col_id_petani = cfg.get("col_id_petani", "")
    col_nama_petani = cfg.get("col_nama_petani", "")
    col_nik = cfg.get("col_nik", "")
    col_jenis_kelamin = cfg.get("col_jenis_kelamin", "")

    if not sheet_name or not col_id_petani or not col_nama_petani or not col_nik or not col_jenis_kelamin:
        raise ValueError(f"missing required fields in params section: {section}")

    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    r0 = row_start - 1
    r1 = row_end

    def c(key: str) -> int:
        val = cfg.get(key, "")
        if not val:
            raise ValueError(f"missing {key} in params section: {section}")
        return _col_letter_to_index(val)

    out = {
        "ID Petani": df.iloc[r0:r1, c("col_id_petani")].astype(str),
        "Nama": df.iloc[r0:r1, c("col_nama_petani")].astype(str),
        "NIK": df.iloc[r0:r1, c("col_nik")].astype(str),
        "Jenis Kelamin": df.iloc[r0:r1, c("col_jenis_kelamin")].astype(str),
    }

    has_bmp = cfg.get("col_bmp_date", "") or cfg.get("col_bmp_name", "")
    has_mk = cfg.get("col_mk_date", "") or cfg.get("col_mk_name", "")
    has_k3 = cfg.get("col_k3_date", "") or cfg.get("col_k3_name", "")

    if has_bmp or has_mk or has_k3:
        out.update(
            {
                "BMP Date": df.iloc[r0:r1, c("col_bmp_date")].astype(str),
                "BMP Name": df.iloc[r0:r1, c("col_bmp_name")].astype(str),
                "BMP Jenis Kelamin": df.iloc[r0:r1, c("col_bmp_jenis_kelamin")].astype(str),
                "BMP Pre Test": df.iloc[r0:r1, c("col_bmp_pre_test")].astype(str),
                "BMP Post Test": df.iloc[r0:r1, c("col_bmp_post_test")].astype(str),
                "BMP Peningkatan": df.iloc[r0:r1, c("col_bmp_peningkatan")].astype(str),
                "MK Date": df.iloc[r0:r1, c("col_mk_date")].astype(str),
                "MK Name": df.iloc[r0:r1, c("col_mk_name")].astype(str),
                "MK Jenis Kelamin": df.iloc[r0:r1, c("col_mk_jenis_kelamin")].astype(str),
                "MK Pre Test": df.iloc[r0:r1, c("col_mk_pre_test")].astype(str),
                "MK Post Test": df.iloc[r0:r1, c("col_mk_post_test")].astype(str),
                "MK Peningkatan": df.iloc[r0:r1, c("col_mk_penigkatan")].astype(str),
                "K3 Date": df.iloc[r0:r1, c("col_k3_date")].astype(str),
                "K3 Name": df.iloc[r0:r1, c("col_k3_name")].astype(str),
                "K3 Jenis Kelamin": df.iloc[r0:r1, c("col_k3_jenis_kelamin")].astype(str),
                "K3 Pre Test": df.iloc[r0:r1, c("col_k3_pre_test")].astype(str),
                "K3 Post Test": df.iloc[r0:r1, c("col_k3_post_test")].astype(str),
                "K3 Peningkatan": df.iloc[r0:r1, c("col_k3_penigkatan")].astype(str),
            }
        )
    else:
        out.update(
            {
                "Training Date": df.iloc[r0:r1, c("col_training_date")].astype(str),
                "Training Name": df.iloc[r0:r1, c("col_training_name")].astype(str),
                "Training Jenis Kelamin": df.iloc[r0:r1, c("col_training_jenis_kelamin")].astype(str),
                "Training Pre Test": df.iloc[r0:r1, c("col_training_pre_test")].astype(str),
                "Training Post Test": df.iloc[r0:r1, c("col_training_post_test")].astype(str),
                "Training Kenaikan": df.iloc[r0:r1, c("col_training_kenaikan")].astype(str),
            }
        )

    out = pd.DataFrame(out)

    out = out[out["ID Petani"].str.strip() != ""]
    out = out.drop_duplicates(subset=["ID Petani"])
    return out


def _apply_table_style(writer: pd.ExcelWriter, sheet_name: str) -> None:
    ws = writer.book[sheet_name]

    # Bold header
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Borders for all used cells
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Autofit columns (approximate)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


def get_flow_status(
    model_id: str = "petani-it6787-kampar-karsem-01",
    csv_path: str = DEFAULT_MODEL_CSV,
) -> List[Tuple[str, str]]:
    """Return step-by-step status messages for the model flow."""
    steps: List[Tuple[str, str]] = []
    try:
        cfg = _load_model_config(model_id, csv_path=csv_path)
    except ModelConfigNotFoundError:
        return [("Failed", "Model config tidak ditemukan")]

    source_gd = cfg.get("source_gd", "").strip()
    if source_gd:
        steps.append(("Done", f"Read Source GD '{source_gd}'"))
    else:
        steps.append(("Failed", "Source GD kosong"))
        return steps

    try:
        _load_params(cfg)
    except Exception as exc:
        steps.append(("Failed", f"Load params: {exc}"))
        return steps

    output_path = _build_output_path(cfg)
    if os.path.exists(output_path):
        steps.append(("Skip", "File sudah ada di input folder"))
        return steps

    source_type = _detect_source_type(source_gd)
    if source_type == "gspread":
        steps.append(("Done", "Detect type: Google Spreadsheet"))
    elif source_type == "excel":
        steps.append(("Done", "Detect type: Excel (.xlsx)"))
    else:
        steps.append(("Failed", "Detect type: Unknown"))

    return steps


def run_flow(
    model_id: str = "petani-it6787-kampar-karsem-01",
    csv_path: str = DEFAULT_MODEL_CSV,
) -> List[Tuple[str, str]]:
    """Run download flow and return step-by-step status messages."""
    steps: List[Tuple[str, str]] = []
    try:
        cfg = _load_model_config(model_id, csv_path=csv_path)
    except ModelConfigNotFoundError:
        return [("Failed", "Model config tidak ditemukan")]

    source_gd = cfg.get("source_gd", "").strip()
    if source_gd:
        steps.append(("Done", f"Read Source GD '{source_gd}'"))
    else:
        steps.append(("Failed", "Source GD kosong"))
        return steps

    try:
        params = _load_params(cfg)
    except Exception as exc:
        steps.append(("Failed", f"Load params: {exc}"))
        return steps

    output_path = _build_output_path(cfg)
    if os.path.exists(output_path):
        steps.append(("Skip", "File sudah ada di input folder"))
        try:
            sheet_list = _get_sheet_list(output_path)
            steps.append(("Done", f"sheet list {sheet_list}"))
            nama_desa_raw = params.get("nama_desa", {}).get("values", "")
            nama_desa = [x.strip() for x in nama_desa_raw.split(",") if x.strip()]
            steps.append(("Done", f"Nama Desa : {', '.join([repr(x) for x in nama_desa])}"))
            df_metadata = _extract_metadata_airterbit(output_path, params)
            steps.append(("Done", "Extract Metadata"))
            df_unique_farmer = _extract_unique_farmer(output_path, params, "unique_farmer")
            steps.append(("Done", "Extract Unique Farmer"))
            df_unique_land = _extract_unique_land_parcel(output_path, params, "unique_land_parcel")
            steps.append(("Done", "Extract Unique Land Parcel"))
            df_training = _extract_training(output_path, params, "training")
            steps.append(("Done", "Extract Training"))
            try:
                output_file = _build_output_file_path(cfg)
                if os.path.exists(output_file):
                    os.remove(output_file)
                    steps.append(("Done", "Output file replaced"))
                with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                    df_metadata.to_excel(writer, index=False, sheet_name="metadata")
                    _apply_table_style(writer, "metadata")
                    df_unique_farmer.to_excel(writer, index=False, sheet_name="unique_farmer")
                    _apply_table_style(writer, "unique_farmer")
                    df_unique_land.to_excel(writer, index=False, sheet_name="unique_land_parcel")
                    _apply_table_style(writer, "unique_land_parcel")
                    df_training.to_excel(writer, index=False, sheet_name="training")
                    _apply_table_style(writer, "training")
                steps.append(("Done", "Output file created"))
            except Exception as exc:
                steps.append(("Failed", f"{exc}"))
        except Exception as exc:
            steps.append(("Failed", f"{exc}"))
        return steps

    source_type = _detect_source_type(source_gd)
    if source_type == "gspread":
        steps.append(("Done", "Detect type: Google Spreadsheet"))
    elif source_type == "excel":
        steps.append(("Done", "Detect type: Excel (.xlsx)"))
    else:
        steps.append(("Failed", "Detect type: Unknown"))
        return steps

    file_id = _extract_gdrive_id(source_gd)
    try:
        if source_type == "gspread":
            _download_google_sheet(file_id, output_path)
        else:
            _download_gdrive_file(file_id, output_path)
        steps.append(("Done", "File Downloaded"))
        sheet_list = _get_sheet_list(output_path)
        steps.append(("Done", f"sheet list {sheet_list}"))
        nama_desa_raw = params.get("nama_desa", {}).get("values", "")
        nama_desa = [x.strip() for x in nama_desa_raw.split(",") if x.strip()]
        steps.append(("Done", f"Nama Desa : {', '.join([repr(x) for x in nama_desa])}"))
        df_metadata = _extract_metadata_airterbit(output_path, params)
        steps.append(("Done", "Extract Metadata"))
        df_unique_farmer = _extract_unique_farmer(output_path, params, "unique_farmer")
        steps.append(("Done", "Extract Unique Farmer"))
        df_unique_land = _extract_unique_land_parcel(output_path, params, "unique_land_parcel")
        steps.append(("Done", "Extract Unique Land Parcel"))
        df_training = _extract_training(output_path, params, "training")
        steps.append(("Done", "Extract Training"))
        try:
            output_file = _build_output_file_path(cfg)
            if os.path.exists(output_file):
                os.remove(output_file)
                steps.append(("Done", "Output file replaced"))
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df_metadata.to_excel(writer, index=False, sheet_name="metadata")
                _apply_table_style(writer, "metadata")
                df_unique_farmer.to_excel(writer, index=False, sheet_name="unique_farmer")
                _apply_table_style(writer, "unique_farmer")
                df_unique_land.to_excel(writer, index=False, sheet_name="unique_land_parcel")
                _apply_table_style(writer, "unique_land_parcel")
                df_training.to_excel(writer, index=False, sheet_name="training")
                _apply_table_style(writer, "training")
            steps.append(("Done", "Output file created"))
        except Exception as exc:
            steps.append(("Failed", f"{exc}"))
    except Exception as exc:
        steps.append(("Failed", f"{exc}"))

    return steps


def download_data(
    model_id: str = "petani-it6787-kampar-karsem-01",
    csv_path: str = DEFAULT_MODEL_CSV,
) -> str:
    """Download Google Drive file for the model and return local path.

    The output path is built from `input_folder` and `input_name` in CSV.
    """
    cfg = _load_model_config(model_id, csv_path=csv_path)

    source_gd = cfg.get("source_gd", "")
    if not source_gd:
        raise ValueError(f"source_gd is empty for model_id: {model_id}")
    output_path = _build_output_path(cfg)
    if os.path.exists(output_path):
        return output_path

    file_id = _extract_gdrive_id(source_gd)
    source_type = _detect_source_type(source_gd)
    if source_type == "gspread":
        _download_google_sheet(file_id, output_path)
    elif source_type == "excel":
        _download_gdrive_file(file_id, output_path)
    else:
        raise RuntimeError("Detect type: Unknown")

    return output_path


if __name__ == "__main__":
    path = download_data()
    print("Downloaded to:", path)
